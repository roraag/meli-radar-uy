"""Exporta la última corrida a Excel (ranking + ofertas crudas).

Salida: outputs/TireShop - Radar ML <fecha>.xlsx
- Hoja "Ranking": una fila por medida, con la brecha como FORMULA VIVA
  (se recalcula si se tocan los precios).
- Hoja "Ofertas": el detalle crudo de la captura.
"""

import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from scoring import ofertas_detalle, ranking

BASE = os.path.dirname(os.path.abspath(__file__))

VERDE = PatternFill("solid", fgColor="C8E6C9")
AMARILLO = PatternFill("solid", fgColor="FFF9C4")
ROJO = PatternFill("solid", fgColor="FFCDD2")
GRIS = PatternFill("solid", fgColor="ECEFF1")
FILL_VEREDICTO = {"barato": VERDE, "alineado": AMARILLO, "caro": ROJO,
                  "sin_datos": GRIS}


def exportar():
    fecha, filas = ranking()
    _, ofertas = ofertas_detalle(fecha)

    wb = Workbook()
    ws = wb.active
    ws.title = "Ranking"

    cols = ["Medida", "Segmento", "Prioridad", "Costo USD", "Precio TireShop",
            "Prom. equiv. USD", "Min equiv.", "Max equiv.", "Brecha %",
            "Veredicto", "Oportunidad", "Ofertas equiv.", "Sellers",
            "% envio gratis", "Premium min", "Mercado tipico jul-26",
            "Cobertura"]
    ws.append(cols)
    for c in ws[1]:
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center")

    for i, r in enumerate(filas, start=2):
        ws.append([
            r["medida"], r["segmento"], r["prioridad"], r["costo_usd"],
            r["precio_final_usd"], r["precio_promedio"], r["precio_min"],
            r["precio_max"], None,  # brecha: formula abajo
            r["veredicto_precio"], r["oportunidad"], r["n_ofertas_equiv"],
            r["sellers_unicos"], r["pct_envio_gratis"], r["premium_min"],
            r["mercado_tipico_jul"],
            "BAJA - validar" if r["cobertura_baja"] else "ok",
        ])
        # brecha % viva: (nuestro - promedio) / promedio
        if r["precio_promedio"]:
            ws.cell(row=i, column=9).value = f"=ROUND(100*(E{i}-F{i})/F{i},1)"
        ws.cell(row=i, column=9).number_format = "0.0"
        ws.cell(row=i, column=10).fill = FILL_VEREDICTO[r["veredicto_precio"]]

    for col, ancho in zip("ABCDEFGHIJKLMNOPQ",
                          (13, 14, 10, 10, 13, 14, 10, 10, 9, 10, 11, 12, 8, 12, 11, 18, 13)):
        ws.column_dimensions[col].width = ancho
    ws.freeze_panes = "A2"

    ws2 = wb.create_sheet("Ofertas")
    ws2.append(["SKU", "Medida", "Clasificacion", "Marca", "Producto",
                "Precio", "Moneda", "Seller", "Envio gratis", "Link catalogo"])
    for c in ws2[1]:
        c.font = Font(bold=True)
    for o in ofertas:
        ws2.append([o[0], o[1], o[2], o[3], o[4], o[5], o[6], o[7],
                    "si" if o[8] else "no",
                    f"https://www.mercadolibre.com.uy/p/{o[9]}"])
    for col, ancho in zip("ABCDEFGHIJ", (11, 12, 12, 12, 50, 9, 8, 12, 11, 42)):
        ws2.column_dimensions[col].width = ancho
    ws2.freeze_panes = "A2"

    os.makedirs(os.path.join(BASE, "outputs"), exist_ok=True)
    path = os.path.join(BASE, "outputs", f"TireShop - Radar ML {fecha}.xlsx")
    wb.save(path)
    print(f"Exportado: {path}")
    return path


if __name__ == "__main__":
    exportar()
